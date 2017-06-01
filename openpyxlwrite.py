import urllib.request
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter

wb = Workbook()

dest_filename = 'empty_book.xlsx'

ws1 = wb.active
ws1.title = "range names"

for row in range (1,40):
	ws1.append(range(600))

ws2 = wb.create_sheet(title="Pi")

ws2['F5'] = 3.14

ws3 = wb.create_sheet(title="Data")
for row in range(1,2):
	for col in range(1,2):
		from bs4 import BeautifulSoup
		req = 'http://www.cnn.com/search?q=' + str(row+col)
		response = urllib.request.urlopen(req)
		soup = BeautifulSoup(response, 'html.parser')
		_ = ws3.cell(column = col + 3, row=row, value = "{0}".format(soup.title.string))
		_ = ws3.cell(column = col, row=row, value = "{0}".format(soup.get_text()))

print(ws3['AA10'].value)

wb.save(filename = dest_filename)